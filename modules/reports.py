import os
import io
from datetime import date, datetime
import pandas as pd
from models import Student, Attendance, Leave, db
from sqlalchemy import extract


def generate_excel_report(month=None, year=None):
    """
    Generate an Excel attendance report.
    Returns: BytesIO object with Excel data.
    """
    today = date.today()
    month = month or today.month
    year = year or today.year

    # Get all active students
    students = Student.query.filter_by(is_active=True).order_by(
        Student.room_number, Student.name
    ).all()

    # Get attendance records for the month
    records = Attendance.query.filter(
        extract('month', Attendance.date) == month,
        extract('year', Attendance.date) == year
    ).all()

    # Build attendance map: {reg_num: {day: status}}
    attendance_map = {}
    for record in records:
        reg = record.registration_number
        day = record.date.day
        if reg not in attendance_map:
            attendance_map[reg] = {}
        attendance_map[reg][day] = record.status

    # Get number of days in month
    import calendar
    num_days = calendar.monthrange(year, month)[1]
    month_name = calendar.month_name[month]

    rows = []
    for student in students:
        reg = student.registration_number
        student_data = attendance_map.get(reg, {})

        present = sum(1 for s in student_data.values() if s == 'Present')
        late = sum(1 for s in student_data.values() if s == 'Late')
        absent = sum(1 for s in student_data.values() if s == 'Absent')
        on_leave = sum(1 for s in student_data.values() if s == 'Leave')
        total_working = num_days
        attendance_pct = round((present + late) / total_working * 100, 1) if total_working > 0 else 0

        row = {
            'Reg. Number': reg,
            'Name': student.name,
            'Room': student.room_number,
            'Department': student.department,
        }

        # Add day columns
        for day in range(1, num_days + 1):
            status = student_data.get(day, '-')
            # Short codes
            if status == 'Present':
                row[str(day)] = 'P'
            elif status == 'Late':
                row[str(day)] = 'L'
            elif status == 'Absent':
                row[str(day)] = 'A'
            elif status == 'Leave':
                row[str(day)] = 'Lv'
            else:
                row[str(day)] = '-'

        row['Present'] = present
        row['Late'] = late
        row['Absent'] = absent
        row['Leave'] = on_leave
        row['Attendance %'] = f"{attendance_pct}%"
        rows.append(row)

    df = pd.DataFrame(rows)

    # Create Excel with formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'{month_name} {year}', index=False)

        workbook = writer.book
        worksheet = writer.sheets[f'{month_name} {year}']

        # Style headers
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Color fills for statuses
        present_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        late_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        absent_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        leave_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

        for col_idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Style data rows
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = center_align
                cell.border = thin_border
                val = cell.value
                if val == 'P':
                    cell.fill = present_fill
                elif val == 'L':
                    cell.fill = late_fill
                elif val == 'A':
                    cell.fill = absent_fill
                elif val == 'Lv':
                    cell.fill = leave_fill

        # Auto-fit columns
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(col)), 6)
            worksheet.column_dimensions[col_letter].width = max_len + 2

        # Add title row
        worksheet.insert_rows(1)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = f'HOSTEL ATTENDANCE REPORT - {month_name.upper()} {year}'
        title_cell.font = Font(bold=True, size=14, color='1E3A5F')
        title_cell.alignment = Alignment(horizontal='center')
        worksheet.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=len(df.columns))

        # Freeze panes
        worksheet.freeze_panes = 'E3'

    output.seek(0)
    return output


def generate_absent_pdf(target_date=None):
    """
    Generate PDF report of absent students.
    Returns: BytesIO object with PDF data.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)

    report_date = target_date or date.today()

    absent_records = Attendance.query.filter_by(
        date=report_date, status='Absent'
    ).all()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                             rightMargin=1.5*cm, leftMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'],
        fontSize=18, textColor=colors.HexColor('#1E3A5F'),
        spaceAfter=10, alignment=1  # Center
    )
    elements.append(Paragraph('HOSTEL NIGHT ATTENDANCE REPORT', title_style))

    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=12, textColor=colors.HexColor('#555555'),
        spaceAfter=5, alignment=1
    )
    elements.append(Paragraph('Absent Students Report', subtitle_style))
    elements.append(Paragraph(f'Date: {report_date.strftime("%d %B %Y")}', subtitle_style))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(HRFlowable(width='100%', thickness=2, color=colors.HexColor('#1E3A5F')))
    elements.append(Spacer(1, 0.5*cm))

    if not absent_records:
        elements.append(Paragraph(
            '✓ No absent students for this date. All students have marked attendance.',
            ParagraphStyle('ok', parent=styles['Normal'],
                           fontSize=12, textColor=colors.green, alignment=1)
        ))
    else:
        elements.append(Paragraph(
            f'Total Absent Students: {len(absent_records)}',
            ParagraphStyle('count', parent=styles['Normal'],
                           fontSize=12, textColor=colors.red,
                           spaceAfter=10, fontName='Helvetica-Bold')
        ))

        # Table data
        table_data = [['#', 'Reg. Number', 'Name', 'Room', 'Department', 'Parent Phone']]

        for idx, record in enumerate(absent_records, 1):
            student = Student.query.filter_by(
                registration_number=record.registration_number
            ).first()
            if student:
                table_data.append([
                    str(idx),
                    student.registration_number,
                    student.name,
                    student.room_number,
                    student.department,
                    student.parent_phone
                ])

        table = Table(table_data, repeatRows=1,
                      colWidths=[1*cm, 3.5*cm, 5*cm, 2*cm, 3.5*cm, 3.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.HexColor('#FFF0F0'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        elements.append(table)

    elements.append(Spacer(1, 1*cm))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.grey))
    elements.append(Spacer(1, 0.3*cm))

    footer_style = ParagraphStyle(
        'Footer', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey, alignment=1
    )
    elements.append(Paragraph(
        f'Generated on {datetime.now().strftime("%d %B %Y at %I:%M %p")} | '
        f'Hostel Attendance Management System',
        footer_style
    ))

    doc.build(elements)
    output.seek(0)
    return output


def get_monthly_stats(month, year):
    """Get monthly attendance statistics."""
    import calendar
    students = Student.query.filter_by(is_active=True).all()
    num_days = calendar.monthrange(year, month)[1]

    stats = []
    for student in students:
        records = Attendance.query.filter_by(
            registration_number=student.registration_number
        ).filter(
            extract('month', Attendance.date) == month,
            extract('year', Attendance.date) == year
        ).all()

        present = sum(1 for r in records if r.status == 'Present')
        late = sum(1 for r in records if r.status == 'Late')
        absent = sum(1 for r in records if r.status == 'Absent')
        on_leave = sum(1 for r in records if r.status == 'Leave')
        pct = round((present + late) / num_days * 100, 1) if num_days > 0 else 0

        stats.append({
            'student': student,
            'present': present,
            'late': late,
            'absent': absent,
            'leave': on_leave,
            'percentage': pct
        })

    return stats
